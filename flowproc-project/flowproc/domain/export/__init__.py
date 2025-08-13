"""
Export domain module for flow cytometry data.
"""

import logging
import pandas as pd
from .service import ExportService
from .excel_writer import ExcelWriter
from .formatters import DataFormatter
from .data_aggregator import DataAggregator
from .replicate_mapper import ReplicateMapper
from .excel_formatter import ExcelFormatter

logger = logging.getLogger(__name__)


# Convenience functions that mimic the old writer API
def process_csv(input_file, output_file, time_course_mode=False, user_replicates=None,
                auto_parse_groups=True, user_group_labels=None, user_groups=None):
    """Process a CSV file to Excel using the export domain services."""
    from pathlib import Path
    from ..parsing import load_and_parse_df, extract_group_animal
    from ..processing.transform import map_replicates
    import logging
    import pandas as pd
    from openpyxl import Workbook
    
    logger = logging.getLogger(__name__)
    
    # Convert paths 
    input_file = Path(input_file)
    output_file = Path(output_file)
    
    # Log processing start (expected by tests)
    logger.info(f"Processing CSV: {input_file}")
    
    # Load and parse the CSV
    df, sid_col = load_and_parse_df(input_file)
    
    # Map replicates
    df, replicate_count = map_replicates(
        df, auto_parse=auto_parse_groups, 
        user_replicates=user_replicates, 
        user_groups=user_groups
    )
    
    if replicate_count == 0:
        logger.warning("No replicates found")
        wb = Workbook()
        wb.remove(wb.active)
        wb.create_sheet("No Data")
        wb.save(output_file)
        logger.info(f"Saved empty output to {output_file}")
        return
    
    # Check if we have time data
    has_time_data = 'Time' in df.columns and df['Time'].notna().any()
    times = sorted(t for t in df["Time"].unique() if pd.notna(t)) if has_time_data else []
    # Treat as timecourse only if there are multiple distinct timepoints
    is_timecourse = len(times) > 1
    
    # Determine processing modes
    # Default: grouped only; if multiple timepoints detected, also generate timecourse
    # If user explicitly selected time_course_mode=True but we do not have multiple timepoints,
    # fall back to grouped-only to avoid creating redundant timecourse files.
    if time_course_mode:
        process_timecourse = is_timecourse
        process_grouped = not is_timecourse  # fallback to grouped when no true timecourse
    else:
        process_grouped = True
        process_timecourse = is_timecourse
    
    logger.info(
        f"Time data detected: {has_time_data} (unique={len(times)}), "
        f"Processing modes - Grouped: {process_grouped}, Timecourse: {process_timecourse}"
    )
    
    # Process in grouped mode if needed
    if process_grouped:
        # Backward-compatible: if caller provided an .xlsx path, use it directly for grouped output
        grouped_output = (
            output_file if output_file.suffix.lower() == ".xlsx"
            else output_file.parent / f"{output_file.stem}_Grouped.xlsx"
        )
        logger.info(f"Processing in grouped mode: {grouped_output}")
        
        # Create workbook for grouped mode
        wb_grouped = Workbook()
        wb_grouped.remove(wb_grouped.active)
        
        # Process and write categories in grouped mode
        process_and_write_categories(
            df, sid_col, wb_grouped, replicate_count, 
            False, user_group_labels  # time_course_mode=False for grouped
        )
        
        # Save grouped workbook
        if not wb_grouped.sheetnames:
            logger.warning(f"No valid data for grouped mode in {input_file}")
            wb_grouped.create_sheet("No Data")
        
        wb_grouped.save(grouped_output)
        logger.info(f"Saved grouped output to {grouped_output}")
    
    # Process in timecourse mode if needed
    if process_timecourse:
        timecourse_output = output_file.parent / f"{output_file.stem}_Timecourse.xlsx"
        logger.info(f"Processing in timecourse mode: {timecourse_output}")
        
        # Create workbook for timecourse mode
        wb_timecourse = Workbook()
        wb_timecourse.remove(wb_timecourse.active)
        
        # Process and write categories in timecourse mode
        process_and_write_categories(
            df, sid_col, wb_timecourse, replicate_count, 
            True, user_group_labels  # time_course_mode=True for timecourse
        )
        
        # Save timecourse workbook
        if not wb_timecourse.sheetnames:
            logger.warning(f"No valid data for timecourse mode in {input_file}")
            wb_timecourse.create_sheet("No Data")
        
        wb_timecourse.save(timecourse_output)
        logger.info(f"Saved timecourse output to {timecourse_output}")
    
    # Only produce the explicit outputs. Do not create a generic "*_Processed.xlsx" copy.
    # If both modes were processed, both _Grouped.xlsx and _Timecourse.xlsx exist.

def process_directory(input_dir, output_dir, recursive=True, pattern="*.csv",
                     status_callback=None, time_course_mode=False, user_replicates=None,
                     auto_parse_groups=True, user_group_labels=None, user_groups=None):
    """Process all CSV files in a directory."""
    from pathlib import Path
    import logging
    
    logger = logging.getLogger(__name__)
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    
    # Find CSV files
    glob_pattern = "**/" + pattern if recursive else pattern
    csv_files = list(input_dir.glob(glob_pattern))
    
    if not csv_files:
        logger.warning(f"No CSV files found in '{input_dir}'")
        if status_callback:
            status_callback("No CSV files found.")
        return 0
    
    count = 0
    for idx, f in enumerate(csv_files, 1):
        if not f.is_file():
            continue
            
        if status_callback:
            status_callback(f"Processing file {idx}/{len(csv_files)}: {f.name}")
        
        # Generate base output filename - the actual output files will be created with suffixes
        base_output = output_dir / f"{f.stem}_Processed"
        
        try:
            process_csv(f, base_output, time_course_mode, user_replicates, 
                       auto_parse_groups, user_group_labels, user_groups)
            count += 1
        except Exception as exc:
            logger.error(f"Error processing '{f}': {exc}")
            if status_callback:
                status_callback(f"Error: {exc}")
    
    if status_callback:
        status_callback(f"Processed {count} files.")
    
    return count

# Constants and keywords from original writer
KEYWORDS = {
    "Freq. of Parent": "freq. of parent",
    "Freq. of Live": "freq. of live",
    "Median": "median",
    "Mean": "mean",
    "Count": "count",
    "GeoMean": "geomean",
    "CV": "cv",
    "SD": "sd",
    "Min": "min",
    "Max": "max",
    "Sum": "sum",
    "Mode": "mode",
    "Range": "range",
}

ALL_TIME_LABEL = "All"

def _clean_column_name(col):
    """Clean column name for processing."""
    return str(col).strip()

def _autofit_columns(worksheet):
    """Auto-fit column widths based on content."""
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell
    
    for col in worksheet.columns:
        max_length = 0
        column_letter = None
        for cell in col:
            if not isinstance(cell, MergedCell):
                column_letter = cell.column_letter
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
        if column_letter:
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

def _get_group_label_map(groups, user_group_labels=None):
    """Get mapping of group numbers to labels."""
    # Convert numpy integers to Python ints
    groups = [int(g) for g in groups]
    if user_group_labels and len(user_group_labels) >= len(groups):
        return {groups[i]: user_group_labels[i] for i in range(len(groups))}
    return {group: f"Group {group}" for group in groups}

def _get_raw_cols(df, sid_col, key_substring):
    """Get columns matching the keyword substring."""
    return [
        c for c in df.columns
        if key_substring in _clean_column_name(c).lower()
        and c not in {sid_col, "Well", "Group", "Animal", "Time", "Replicate", "Tissue"}
    ]

def _create_sheet_pair(wb, sheet_root, num_replicates, raw_cols, group_label_map, groups, tissues_detected, is_time_course=False, has_time_data=False):
    """Create a pair of worksheets (values and IDs) with headers."""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    
    # Ensure sheet names don't exceed Excel's 31 character limit
    ws_vals = wb.create_sheet(sheet_root[:31])
    ws_ids = wb.create_sheet(f"{sheet_root} IDs"[:31])
    
    # For timecourse mode, headers are handled in _write_timecourse_data
    if is_time_course:
        return ws_vals, ws_ids
    
    # Determine column offset and headers for non-timecourse modes
    if tissues_detected and has_time_data:
        col_offset = 4  # Group, Time, Tissue, then data
        # Merge Group, Time, and Tissue headers across rows 1-2
        for ws in (ws_vals, ws_ids):
            ws.merge_cells("A1:A2")
            ws.merge_cells("B1:B2")
            ws.merge_cells("C1:C2")
            ws.cell(row=1, column=1, value="Group").alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=1, column=2, value="Time").alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=1, column=3, value="Tissue").alignment = Alignment(horizontal="center", vertical="center")
    elif tissues_detected:
        col_offset = 3  # Group, Tissue, then data
        # Merge Group and Tissue headers across rows 1-2
        for ws in (ws_vals, ws_ids):
            ws.merge_cells("A1:A2")
            ws.merge_cells("B1:B2")
            ws.cell(row=1, column=1, value="Group").alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=1, column=2, value="Tissue").alignment = Alignment(horizontal="center", vertical="center")
    elif has_time_data:
        col_offset = 3  # Group, Time, then data
        # Merge Group and Time headers across rows 1-2
        for ws in (ws_vals, ws_ids):
            ws.merge_cells("A1:A2")
            ws.merge_cells("B1:B2")
            ws.cell(row=1, column=1, value="Group").alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=1, column=2, value="Time").alignment = Alignment(horizontal="center", vertical="center")
    else:
        col_offset = 2  # Group, then data
        # Merge Group header across rows 1-2
        for ws in (ws_vals, ws_ids):
            ws.merge_cells("A1:A2")
            ws.cell(row=1, column=1, value="Group").alignment = Alignment(horizontal="center", vertical="center")

    # Add headers for each metric column
    for idx, col_name in enumerate(raw_cols, start=1):
        start_col = col_offset + (idx - 1) * num_replicates
        end_col = col_offset + idx * num_replicates - 1
        
        # Add merged header for the metric
        for ws in (ws_vals, ws_ids):
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)
            ws.merge_cells(f"{start_letter}1:{end_letter}1")
            ws.cell(row=1, column=start_col, value=col_name).alignment = Alignment(horizontal="center")
            
            # Add replicate headers
            for rep in range(num_replicates):
                ws.cell(row=2, column=start_col + rep, value=f"Rep {rep + 1}")
    
    return ws_vals, ws_ids

def _process_metric(df, sid_col, wb, num_replicates, cat, key_substring, time_course_mode, tissues_detected, group_label_map, times, groups):
    """Process a single metric category and create sheets."""
    raw_cols = _get_raw_cols(df, sid_col, key_substring)
    if not raw_cols:
        logger.info(f"No columns for '{cat}'")
        return
    
    # Check if we have time data
    has_time_data = 'Time' in df.columns and df['Time'].notna().any()
    
    sheet_root = cat[:31]
    ws_vals, ws_ids = _create_sheet_pair(
        wb, sheet_root, num_replicates, raw_cols, 
        group_label_map, groups, tissues_detected, time_course_mode, has_time_data
    )

    if time_course_mode:
        _write_timecourse_data(
            df, sid_col, ws_vals, ws_ids, raw_cols, 
            num_replicates, times, groups, group_label_map
        )
    else:
        _write_standard_data(
            df, sid_col, ws_vals, ws_ids, raw_cols, 
            num_replicates, times, groups, group_label_map, tissues_detected
        )

    _autofit_columns(ws_vals)
    _autofit_columns(ws_ids)

def _write_standard_data(df, sid_col, ws_vals, ws_ids, raw_cols, num_replicates, times, groups, group_label_map, tissues_detected):
    """Write data in standard (non-time-course) format."""
    from ..processing.transform import reshape_pair
    from ..parsing import get_tissue_full_name
    
    # Check if we have time data
    has_time_data = 'Time' in df.columns and df['Time'].notna().any()
    
    # Adjust column offset based on what columns we need
    if tissues_detected and has_time_data:
        col_offset = 4  # Group, Time, Tissue, then data
    elif tissues_detected or has_time_data:
        col_offset = 3  # Group + (Time OR Tissue), then data
    else:
        col_offset = 2  # Just Group, then data
    
    # Process each column separately
    all_data = []
    
    for col_idx, col in enumerate(raw_cols):
        # Get blocks for this specific column
        val_blocks, id_blocks, tissue_codes, group_numbers, time_values = reshape_pair(
            df, sid_col, [col], num_replicates, use_tissue=tissues_detected, include_time=has_time_data
        )
        
        if val_blocks:
            all_data.append((col_idx, col, val_blocks, id_blocks, time_values))
            
            # Write group, time, and tissue info (only once per row)
            if col_idx == 0:
                for row_idx, (group, tissue_code, time_val) in enumerate(zip(group_numbers, tissue_codes, time_values), start=3):
                    ws_vals.cell(row=row_idx, column=1, value=group_label_map.get(int(group), f"Group {group}"))
                    ws_ids.cell(row=row_idx, column=1, value=group_label_map.get(int(group), f"Group {group}"))
                    
                    current_col = 2
                    
                    if has_time_data:
                        time_str = _format_time(time_val) if time_val is not None else ""
                        ws_vals.cell(row=row_idx, column=current_col, value=time_str)
                        ws_ids.cell(row=row_idx, column=current_col, value=time_str)
                        current_col += 1
                    
                    if tissues_detected:
                        tissue = tissue_code[0] if isinstance(tissue_code, tuple) else tissue_code
                        tissue_name = get_tissue_full_name(tissue)
                        ws_vals.cell(row=row_idx, column=current_col, value=tissue_name)
                        ws_ids.cell(row=row_idx, column=current_col, value=tissue_name)
    
    # Write data blocks
    for col_idx, col_name, val_blocks, id_blocks, _ in all_data:
        start_col = col_offset + col_idx * num_replicates
        
        for row_idx, (val_row, id_row) in enumerate(zip(val_blocks, id_blocks), start=3):
            for rep_idx, (val, sid) in enumerate(zip(val_row, id_row)):
                col = start_col + rep_idx
                ws_vals.cell(row=row_idx, column=col, value=val if pd.notna(val) else None)
                ws_ids.cell(row=row_idx, column=col, value=sid)

def _write_timecourse_data(df, sid_col, ws_vals, ws_ids, raw_cols, num_replicates, times, groups, group_label_map):
    """Write data in time-course format with times in columns."""
    # Structure: Row 1 = Population, Row 2 = Groups (no gaps), Row 3 = Time label + Replicates, Column A = Time points
    # Empty column gap between different populations (truly empty, no "None" values)
    
    # Write "Time" label in Column A and merge across rows 1-3 for clean look
    ws_vals.cell(row=1, column=1, value="Time")
    ws_vals.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    ws_ids.cell(row=1, column=1, value="Time")
    ws_ids.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    
    # Write time labels in Column A (starting from row 4)
    for row_idx, time_val in enumerate(times, start=4):
        time_str = _format_time(time_val)
        ws_vals.cell(row=row_idx, column=1, value=time_str)
        ws_ids.cell(row=row_idx, column=1, value=time_str)
    
    # Calculate total columns needed for this population
    # Each group gets num_replicates columns, no gaps between groups
    total_cols = len(groups) * num_replicates
    
    # For each population/metric, create the structure
    for pop_idx, col_name in enumerate(raw_cols):
        # Calculate starting column for this population
        # Add 1 extra column gap between populations (except for the first one)
        pop_start_col = 2 + (pop_idx * (total_cols + 1))
        pop_end_col = pop_start_col + total_cols - 1
        
        # Row 1: Population name (merge across ALL columns for this population)
        ws_vals.merge_cells(start_row=1, start_column=pop_start_col, end_row=1, end_column=pop_end_col)
        ws_vals.cell(row=1, column=pop_start_col, value=col_name)
        ws_ids.merge_cells(start_row=1, start_column=pop_start_col, end_row=1, end_column=pop_end_col)
        ws_ids.cell(row=1, column=pop_start_col, value=col_name)
        
        # Row 2: Group headers (each group gets its own merged section)
        current_col = pop_start_col
        for group_idx, group in enumerate(groups):
            # Group header (merge across replicates)
            group_start_col = current_col
            group_end_col = current_col + num_replicates - 1
            
            # Merge cells for this group across its replicate columns
            ws_vals.merge_cells(start_row=2, start_column=group_start_col, end_row=2, end_column=group_end_col)
            ws_vals.cell(row=2, column=group_start_col, value=f"Group {group}")
            ws_ids.merge_cells(start_row=2, start_column=group_start_col, end_row=2, end_column=group_end_col)
            ws_ids.cell(row=2, column=group_start_col, value=f"Group {group}")
            
            # Replicate headers under each group (in Row 3)
            for rep in range(1, num_replicates + 1):
                col_pos = group_start_col + rep - 1
                ws_vals.cell(row=3, column=col_pos, value=f"Replicate {rep}")
                ws_ids.cell(row=3, column=col_pos, value=f"Replicate {rep}")
            
            # Move to next group position (no gaps between groups)
            current_col = group_end_col + 1
        
        # Data rows: Write actual values for each time point
        for time_idx, time_val in enumerate(times):
            data_row = 4 + time_idx
            
            # Write data for each group and replicate at this time point
            current_col = pop_start_col
            for group in groups:
                for rep in range(1, num_replicates + 1):
                    col_pos = current_col + rep - 1
                    
                    # Find the data for this specific group, replicate, and time
                    subset = df[(df["Group"] == group) & (df["Replicate"] == rep) & (df["Time"] == time_val)]
                    
                    if not subset.empty and col_name in subset.columns:
                        val = subset[col_name].iloc[0]
                        sid = subset[sid_col].iloc[0]
                        ws_vals.cell(row=data_row, column=col_pos, value=val if pd.notna(val) else None)
                        ws_ids.cell(row=data_row, column=col_pos, value=sid)
                    else:
                        # Don't write anything for missing data - leave cell empty
                        pass
                
                # Move to next group position (no gaps between groups)
                current_col = current_col + num_replicates

def _format_time(t):
    """Format time value for display."""
    if t is None:
        return ALL_TIME_LABEL
    hours = int(t)
    minutes = int((t - hours) * 60)
    return f"{hours}:{minutes:02d}"

def process_and_write_categories(df, sid_col, wb, n, time_course_mode=False, user_group_labels=None):
    """Create one pair of sheets per metric (e.g., Count, Median)."""
    if df.empty:
        logger.warning("Empty DataFrame; creating 'No Data' sheet")
        wb.create_sheet("No Data")
        return
    
    # Get times and groups
    times = sorted(t for t in df["Time"].unique() if pd.notna(t)) if "Time" in df.columns and df["Time"].notna().any() else [None]
    groups = sorted(df["Group"].unique())
    
    # Convert numpy types to Python types
    groups = [int(g) for g in groups]
    group_label_map = _get_group_label_map(groups, user_group_labels)

    logger.debug(f"Replicates: {n}, Times: {times}, Groups: {groups}, Labels: {group_label_map}")
    
    # Check if we have time data
    is_timecourse = len(times) > 1 or (len(times) == 1 and times[0] is not None)
    if time_course_mode and not is_timecourse:
        logger.warning("Time-course mode selected but no time data; falling back.")
        time_course_mode = False

    # Check if we have tissues (any tissue, not just multiple)
    from ..parsing import extract_tissue
    tissue_codes = df[sid_col].apply(lambda x: extract_tissue(str(x)) if pd.notna(x) else 'UNK')
    tissues_detected = (tissue_codes != 'UNK').any()

    # Process each metric category
    for cat, key_substring in KEYWORDS.items():
        _process_metric(
            df, sid_col, wb, n, cat, key_substring, 
            time_course_mode, tissues_detected, group_label_map, times, groups
        )

def _create_empty_excel(output_file, sheet_name):
    """Create an empty Excel file with given sheet name."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet(sheet_name)
    wb.save(output_file)




__all__ = [
    'ExportService',
    'ExcelWriter', 
    'DataAggregator',
    'ReplicateMapper',
    'ExcelFormatter',
    'process_csv',
    'process_directory'
] 
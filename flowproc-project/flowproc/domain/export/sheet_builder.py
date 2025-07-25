"""Build Excel sheets with proper structure."""
from typing import List, Dict, Optional, Tuple, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
import logging

from .style_manager import StyleManager
from .time_formatter import TimeFormatter

logger = logging.getLogger(__name__)


class SheetBuilder:
    """Builds structured Excel sheets."""
    
    def __init__(self,
                 style_manager: Optional[StyleManager] = None,
                 time_formatter: Optional[TimeFormatter] = None):
        """
        Initialize sheet builder.
        
        Args:
            style_manager: Style manager instance
            time_formatter: Time formatter instance
        """
        self.style_manager = style_manager or StyleManager()
        self.time_formatter = time_formatter or TimeFormatter()
        
    def create_data_sheet(self, wb: Workbook,
                         sheet_name: str,
                         data: pd.DataFrame,
                         metadata_cols: List[str],
                         value_cols: List[str],
                         n_replicates: int) -> Worksheet:
        """
        Create a data sheet with values and formatting.
        
        Args:
            wb: Workbook
            sheet_name: Name for the sheet
            data: DataFrame with data
            metadata_cols: Metadata column names (Group, Time, etc.)
            value_cols: Value column names
            n_replicates: Number of replicates
            
        Returns:
            Created worksheet
        """
        # Ensure sheet name is valid (Excel limit: 31 chars)
        sheet_name = sheet_name[:31]
        
        ws = wb.create_sheet(sheet_name)
        
        # Write headers
        self._write_headers(ws, metadata_cols, value_cols, n_replicates)
        
        # Write data
        self._write_data(ws, data, metadata_cols, value_cols)
        
        # Apply formatting
        self._format_sheet(ws, metadata_cols, value_cols)
        
        return ws
        
    def create_id_sheet(self, wb: Workbook,
                       sheet_name: str,
                       data: pd.DataFrame,
                       metadata_cols: List[str],
                       value_cols: List[str],
                       n_replicates: int) -> Worksheet:
        """
        Create an ID sheet showing sample IDs.
        
        Args:
            wb: Workbook
            sheet_name: Name for the sheet
            data: DataFrame with data and sample IDs
            metadata_cols: Metadata column names
            value_cols: Value column names
            n_replicates: Number of replicates
            
        Returns:
            Created worksheet
        """
        # Ensure sheet name is valid
        id_sheet_name = f"{sheet_name[:28]} ID"[:31]
        
        ws = wb.create_sheet(id_sheet_name)
        
        # Write headers (same as data sheet)
        self._write_headers(ws, metadata_cols, value_cols, n_replicates)
        
        # Write sample IDs instead of values
        self._write_ids(ws, data, metadata_cols, value_cols)
        
        # Apply formatting
        self._format_sheet(ws, metadata_cols, value_cols, is_id_sheet=True)
        
        return ws
        
    def _write_headers(self, ws: Worksheet,
                      metadata_cols: List[str],
                      value_cols: List[str],
                      n_replicates: int) -> None:
        """Write sheet headers."""
        col_offset = len(metadata_cols)
        
        # Write metadata headers (merged across rows 1-2)
        for i, col_name in enumerate(metadata_cols, 1):
            ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)
            ws.cell(row=1, column=i, value=col_name)
            
        # Write value column headers with merged cells
        for i, col_name in enumerate(value_cols):
            start_col = col_offset + i * n_replicates + 1
            end_col = start_col + n_replicates - 1
            
            # Merge cells for column name
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col
            )
            ws.cell(row=1, column=start_col, value=col_name)
            
            # Write replicate headers
            for j in range(n_replicates):
                ws.cell(row=2, column=start_col + j, value=f"Rep {j + 1}")
                
        # Apply header styling
        self.style_manager.apply_header_style(ws, row=1)
        self.style_manager.apply_subheader_style(ws, row=2)
        
    def _write_data(self, ws: Worksheet,
                   data: pd.DataFrame,
                   metadata_cols: List[str],
                   value_cols: List[str]) -> None:
        """Write data values to sheet."""
        row_offset = 3  # After headers
        
        for idx, (_, row) in enumerate(data.iterrows()):
            excel_row = idx + row_offset
            
            # Write metadata
            for col_idx, col_name in enumerate(metadata_cols, 1):
                value = row.get(col_name, '')
                
                # Format time values
                if col_name == 'Time' and value is not None and value != '':
                    value = self.time_formatter.format(value)
                    
                ws.cell(row=excel_row, column=col_idx, value=value)
                
            # Write values
            col_offset = len(metadata_cols)
            for col_name in value_cols:
                if col_name in row:
                    value = row[col_name]
                    
                    # Handle different value types
                    if pd.isna(value):
                        display_value = None
                    elif isinstance(value, (list, pd.Series)):
                        # Multiple replicate values
                        for i, val in enumerate(value):
                            if i < len(value):
                                ws.cell(
                                    row=excel_row,
                                    column=col_offset + i + 1,
                                    value=val if not pd.isna(val) else None
                                )
                    else:
                        # Single value
                        ws.cell(
                            row=excel_row,
                            column=col_offset + 1,
                            value=value
                        )
                        
                col_offset += len(row.get(col_name, [None]))
                
    def _write_ids(self, ws: Worksheet,
                  data: pd.DataFrame,
                  metadata_cols: List[str],
                  value_cols: List[str]) -> None:
        """Write sample IDs to sheet."""
        row_offset = 3
        
        for idx, (_, row) in enumerate(data.iterrows()):
            excel_row = idx + row_offset
            
            # Write metadata (same as data sheet)
            for col_idx, col_name in enumerate(metadata_cols, 1):
                value = row.get(col_name, '')
                
                if col_name == 'Time' and value is not None and value != '':
                    value = self.time_formatter.format(value)
                    
                ws.cell(row=excel_row, column=col_idx, value=value)
                
            # Write sample IDs
            sample_id = row.get('SampleID', '')
            col_offset = len(metadata_cols)
            
            for col_name in value_cols:
                if col_name in row:
                    # Write same ID for all replicates of this column
                    n_reps = len(row.get(col_name, [None]))
                    for i in range(n_reps):
                        ws.cell(
                            row=excel_row,
                            column=col_offset + i + 1,
                            value=sample_id
                        )
                        
                col_offset += n_reps
                
    def _format_sheet(self, ws: Worksheet,
                     metadata_cols: List[str],
                     value_cols: List[str],
                     is_id_sheet: bool = False) -> None:
        """Apply formatting to sheet."""
        # Format columns by type
        if not is_id_sheet:
            col_offset = len(metadata_cols)
            for i, col_name in enumerate(value_cols):
                start_col = col_offset + i * 3 + 1  # Assuming 3 replicates
                
                # Format all replicate columns
                for j in range(3):
                    self.style_manager.format_column_by_type(
                        ws, start_col + j, col_name
                    )
                    
        # Apply borders
        if ws.max_row > 2:
            self.style_manager.apply_borders(
                ws, start_row=1, end_row=ws.max_row
            )
            
        # Apply alternating rows
        if ws.max_row > 3:
            self.style_manager.apply_alternating_rows(
                ws, start_row=3, end_row=ws.max_row
            )
            
        # Auto-fit columns
        self.style_manager.autofit_columns(ws)
        
        # Freeze panes
        self.style_manager.freeze_panes(ws, row=3, col=len(metadata_cols) + 1)
        
    def create_summary_sheet(self, wb: Workbook,
                            summary_data: pd.DataFrame) -> Worksheet:
        """
        Create a summary sheet with aggregated data.
        
        Args:
            wb: Workbook
            summary_data: Aggregated summary data
            
        Returns:
            Created worksheet
        """
        ws = wb.create_sheet("Summary")
        
        # Write data starting from A1
        for r_idx, row in enumerate(summary_data.itertuples(index=False), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
                
        # Format headers
        if len(summary_data) > 0:
            # Write column headers
            for c_idx, col_name in enumerate(summary_data.columns, 1):
                ws.cell(row=1, column=c_idx, value=col_name)
                
            self.style_manager.apply_header_style(ws, row=1)
            
            # Apply formatting
            self.style_manager.apply_borders(
                ws, start_row=1, end_row=len(summary_data) + 1
            )
            self.style_manager.autofit_columns(ws)
            self.style_manager.freeze_panes(ws, row=2)
            
        return ws
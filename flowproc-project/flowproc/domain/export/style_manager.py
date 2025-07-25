"""Manage Excel styles and formatting."""
from typing import Dict, Optional
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle
)
from openpyxl.worksheet.worksheet import Worksheet
import logging

logger = logging.getLogger(__name__)


class StyleManager:
    """Manages Excel styles and formatting."""
    
    def __init__(self):
        """Initialize style manager with default styles."""
        self._init_styles()
        
    def _init_styles(self) -> None:
        """Initialize default styles."""
        # Fonts
        self.header_font = Font(bold=True, size=12)
        self.subheader_font = Font(bold=True, size=10)
        self.normal_font = Font(size=10)
        
        # Fills
        self.header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        self.subheader_fill = PatternFill(
            start_color="DCE6F1",
            end_color="DCE6F1",
            fill_type="solid"
        )
        self.alternating_fill = PatternFill(
            start_color="F2F2F2",
            end_color="F2F2F2",
            fill_type="solid"
        )
        
        # Alignment
        self.center_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        self.left_alignment = Alignment(
            horizontal="left",
            vertical="center"
        )
        
        # Borders
        thin_side = Side(style="thin", color="000000")
        self.thin_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )
        
        # Named styles
        self._create_named_styles()
        
    def _create_named_styles(self) -> None:
        """Create named styles for common formats."""
        # Percentage style
        self.percentage_style = NamedStyle(name="percentage")
        self.percentage_style.number_format = '0.00%'
        self.percentage_style.alignment = self.center_alignment
        
        # Number style
        self.number_style = NamedStyle(name="number")
        self.number_style.number_format = '0.00'
        self.number_style.alignment = self.center_alignment
        
        # Integer style
        self.integer_style = NamedStyle(name="integer")
        self.integer_style.number_format = '0'
        self.integer_style.alignment = self.center_alignment
        
    def apply_header_style(self, ws: Worksheet, row: int,
                          start_col: int = 1, end_col: Optional[int] = None) -> None:
        """
        Apply header style to a row.
        
        Args:
            ws: Worksheet
            row: Row number
            start_col: Starting column
            end_col: Ending column (inclusive)
        """
        if end_col is None:
            end_col = ws.max_column
            
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
            
            # White text on dark background
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            
    def apply_subheader_style(self, ws: Worksheet, row: int,
                             start_col: int = 1, end_col: Optional[int] = None) -> None:
        """Apply subheader style to a row."""
        if end_col is None:
            end_col = ws.max_column
            
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
            
    def apply_alternating_rows(self, ws: Worksheet,
                              start_row: int, end_row: int,
                              start_col: int = 1, end_col: Optional[int] = None) -> None:
        """Apply alternating row colors."""
        if end_col is None:
            end_col = ws.max_column
            
        for row in range(start_row, end_row + 1):
            if (row - start_row) % 2 == 1:  # Odd rows (0-indexed)
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = self.alternating_fill
                    
    def apply_borders(self, ws: Worksheet,
                     start_row: int, end_row: int,
                     start_col: int = 1, end_col: Optional[int] = None) -> None:
        """Apply borders to a range."""
        if end_col is None:
            end_col = ws.max_column
            
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
                
    def format_column_by_type(self, ws: Worksheet, col: int,
                             col_name: str, start_row: int = 3) -> None:
        """
        Format column based on data type.
        
        Args:
            ws: Worksheet
            col: Column number
            col_name: Column name for type detection
            start_row: First data row
        """
        col_lower = col_name.lower()
        
        # Determine format based on column name
        if 'freq' in col_lower or '%' in col_name:
            number_format = '0.00%'
            divide_by_100 = '%' not in col_name  # Divide by 100 if not already percentage
        elif 'count' in col_lower:
            number_format = '#,##0'
            divide_by_100 = False
        elif any(term in col_lower for term in ['mean', 'median', 'cv', 'sd']):
            number_format = '#,##0.00'
            divide_by_100 = False
        else:
            number_format = 'General'
            divide_by_100 = False
            
        # Apply format to all cells in column
        for row in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            
            if divide_by_100 and cell.value is not None:
                try:
                    cell.value = float(cell.value) / 100
                except (TypeError, ValueError):
                    pass
                    
            cell.number_format = number_format
            cell.alignment = self.center_alignment
            
    def autofit_columns(self, ws: Worksheet, min_width: int = 8,
                       max_width: int = 50) -> None:
        """
        Auto-fit column widths based on content.
        
        Args:
            ws: Worksheet
            min_width: Minimum column width
            max_width: Maximum column width
        """
        for column in ws.columns:
            max_length = 0
            column_letter = None
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                        column_letter = cell.column_letter
                except:
                    pass
                    
            if column_letter:
                adjusted_width = min(max(max_length + 2, min_width), max_width)
                ws.column_dimensions[column_letter].width = adjusted_width
                
    def freeze_panes(self, ws: Worksheet, row: int = 3, col: int = 1) -> None:
        """
        Freeze panes for easier viewing.
        
        Args:
            ws: Worksheet
            row: Row below which to freeze
            col: Column to the right of which to freeze
        """
        from openpyxl.utils import get_column_letter
        freeze_cell = f"{get_column_letter(col)}{row}"
        ws.freeze_panes = freeze_cell
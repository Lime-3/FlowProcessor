"""Format Excel workbooks with consistent styling."""
from typing import Optional, List, Dict, Any
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import logging

from .style_manager import StyleManager

logger = logging.getLogger(__name__)


class ExcelFormatter:
    """Applies consistent formatting to Excel workbooks."""
    
    def __init__(self, style_manager: Optional[StyleManager] = None):
        """
        Initialize Excel formatter.
        
        Args:
            style_manager: Style manager instance
        """
        self.style_manager = style_manager or StyleManager()
        
    def format_workbook(self, wb: Workbook,
                       sheet_order: Optional[List[str]] = None) -> None:
        """
        Apply formatting to entire workbook.
        
        Args:
            wb: Workbook to format
            sheet_order: Desired sheet order
        """
        # Remove default sheet if empty
        if 'Sheet' in wb.sheetnames:
            default_sheet = wb['Sheet']
            if default_sheet.max_row == 1 and default_sheet.max_column == 1:
                wb.remove(default_sheet)
                
        # Reorder sheets if specified
        if sheet_order:
            self._reorder_sheets(wb, sheet_order)
            
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self._format_worksheet(ws)
            
        # Set active sheet
        if wb.sheetnames:
            wb.active = wb[wb.sheetnames[0]]
            
    def _reorder_sheets(self, wb: Workbook, sheet_order: List[str]) -> None:
        """Reorder sheets in workbook."""
        current_sheets = wb.sheetnames.copy()
        
        # Create new order
        new_order = []
        
        # Add sheets in specified order
        for name in sheet_order:
            if name in current_sheets:
                new_order.append(name)
                current_sheets.remove(name)
                
        # Add remaining sheets
        new_order.extend(current_sheets)
        
        # Reorder by moving sheets
        for idx, sheet_name in enumerate(new_order):
            if sheet_name in wb.sheetnames:
                wb.move_sheet(sheet_name, idx - wb.sheetnames.index(sheet_name))
                
    def _format_worksheet(self, ws: Worksheet) -> None:
        """Apply formatting to a worksheet."""
        if ws.max_row <= 1:
            return  # Empty sheet
            
        # Auto-fit columns
        self.style_manager.autofit_columns(ws)
        
        # Ensure borders are complete
        if ws.max_row > 1:
            self.style_manager.apply_borders(
                ws, start_row=1, end_row=ws.max_row
            )
            
    def add_info_sheet(self, wb: Workbook,
                      processing_info: Dict[str, Any]) -> None:
        """
        Add information sheet with processing details.
        
        Args:
            wb: Workbook
            processing_info: Dictionary with processing information
        """
        ws = wb.create_sheet("Info", 0)  # Insert at beginning
        
        # Add title
        ws.cell(row=1, column=1, value="Flow Cytometry Processing Information")
        ws.merge_cells('A1:B1')
        self.style_manager.apply_header_style(ws, row=1, end_col=2)
        
        # Add information
        row = 3
        info_items = [
            ("Processing Date", processing_info.get('date', 'N/A')),
            ("Software Version", processing_info.get('version', 'N/A')),
            ("Input Files", processing_info.get('input_files', 'N/A')),
            ("Output Mode", processing_info.get('mode', 'Standard')),
            ("Groups", processing_info.get('groups', 'Auto-detected')),
            ("Replicates", processing_info.get('replicates', 'Auto-detected')),
            ("Total Samples", processing_info.get('total_samples', 0)),
            ("Valid Samples", processing_info.get('valid_samples', 0)),
        ]
        
        for label, value in info_items:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
            
        # Format
        self.style_manager.autofit_columns(ws)
        self.style_manager.apply_borders(ws, start_row=3, end_row=row-1, end_col=2)
        
    def add_validation_sheet(self, wb: Workbook,
                           validation_results: List[str]) -> None:
        """
        Add validation results sheet.
        
        Args:
            wb: Workbook
            validation_results: List of validation messages
        """
        if not validation_results:
            return
            
        ws = wb.create_sheet("Validation")
        
        # Add title
        ws.cell(row=1, column=1, value="Data Validation Results")
        self.style_manager.apply_header_style(ws, row=1)
        
        # Add messages
        for idx, message in enumerate(validation_results, 3):
            ws.cell(row=idx, column=1, value=message)
            
        # Format
        self.style_manager.autofit_columns(ws)
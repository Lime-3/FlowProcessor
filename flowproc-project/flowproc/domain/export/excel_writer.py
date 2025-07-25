"""
Excel writer for flow cytometry data export.
"""

from typing import Dict, List, Any, Optional
import pandas as pd
from pathlib import Path
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Handles Excel file writing for flow cytometry data."""
    
    def __init__(self):
        """Initialize the Excel writer."""
        self.default_styles = {
            'header_font': Font(bold=True, color='FFFFFF'),
            'header_fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'header_alignment': Alignment(horizontal='center', vertical='center'),
            'data_font': Font(),
            'data_alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    def write_excel(self, dataframes: List[pd.DataFrame], filepath: str,
                   sheet_names: Optional[List[str]] = None,
                   include_index: bool = False,
                   auto_adjust_columns: bool = True) -> None:
        """
        Write DataFrames to Excel file.
        
        Args:
            dataframes: List of DataFrames to write
            filepath: Output file path
            sheet_names: Names for each sheet (optional)
            include_index: Whether to include DataFrame index
            auto_adjust_columns: Whether to auto-adjust column widths
        """
        try:
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Write each DataFrame to a sheet
            for i, df in enumerate(dataframes):
                sheet_name = sheet_names[i] if sheet_names and i < len(sheet_names) else f'Sheet{i+1}'
                sheet_name = self._sanitize_sheet_name(sheet_name)
                
                # Create worksheet
                ws = wb.create_sheet(title=sheet_name)
                
                # Write data
                self._write_dataframe_to_sheet(ws, df, include_index)
                
                # Apply formatting
                self._apply_sheet_formatting(ws, df, include_index)
                
                # Auto-adjust columns
                if auto_adjust_columns:
                    self._auto_adjust_columns(ws, df, include_index)
            
            # Save workbook
            wb.save(filepath)
            logger.info(f"Excel file saved: {filepath}")
            
        except Exception as e:
            logger.error(f"Failed to write Excel file: {e}")
            raise
    
    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame, include_index: bool) -> None:
        """Write DataFrame to worksheet."""
        # Convert DataFrame to rows
        rows = dataframe_to_rows(df, index=include_index, header=True)
        
        # Write rows to worksheet
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    
    def _apply_sheet_formatting(self, ws, df: pd.DataFrame, include_index: bool) -> None:
        """Apply formatting to worksheet."""
        # Get header row range
        header_row = 1
        start_col = 2 if include_index else 1
        end_col = len(df.columns) + (1 if include_index else 0)
        
        # Format header row
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = self.default_styles['header_font']
            cell.fill = self.default_styles['header_fill']
            cell.alignment = self.default_styles['header_alignment']
            cell.border = self.default_styles['border']
        
        # Format data rows
        for row in range(2, len(df) + 2):
            for col in range(1, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = self.default_styles['data_font']
                cell.alignment = self.default_styles['data_alignment']
                cell.border = self.default_styles['border']
    
    def _auto_adjust_columns(self, ws, df: pd.DataFrame, include_index: bool) -> None:
        """Auto-adjust column widths."""
        # Calculate column widths
        column_widths = []
        
        # Add index column width if needed
        if include_index:
            column_widths.append(15)  # Default index width
        
        # Calculate widths for data columns
        for col in df.columns:
            # Get maximum length in column
            max_length = max(
                len(str(col)),  # Header length
                df[col].astype(str).str.len().max() if len(df) > 0 else 0  # Data length
            )
            # Add some padding
            column_widths.append(min(max_length + 2, 50))
        
        # Apply column widths
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize sheet name for Excel compatibility."""
        # Excel sheet names have restrictions
        invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
        for char in invalid_chars:
            name = name.replace(char, '_')
        
        # Limit length to 31 characters
        if len(name) > 31:
            name = name[:31]
        
        return name
    
    def write_with_metadata(self, dataframes: List[pd.DataFrame], filepath: str,
                           metadata: Dict[str, Any],
                           sheet_names: Optional[List[str]] = None) -> None:
        """
        Write DataFrames to Excel with metadata sheet.
        
        Args:
            dataframes: List of DataFrames to write
            filepath: Output file path
            metadata: Metadata to include in info sheet
            sheet_names: Names for each sheet (optional)
        """
        try:
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Add metadata sheet
            self._add_metadata_sheet(wb, metadata)
            
            # Write data sheets
            for i, df in enumerate(dataframes):
                sheet_name = sheet_names[i] if sheet_names and i < len(sheet_names) else f'Data_{i+1}'
                sheet_name = self._sanitize_sheet_name(sheet_name)
                
                ws = wb.create_sheet(title=sheet_name)
                self._write_dataframe_to_sheet(ws, df, include_index=False)
                self._apply_sheet_formatting(ws, df, include_index=False)
                self._auto_adjust_columns(ws, df, include_index=False)
            
            # Save workbook
            wb.save(filepath)
            logger.info(f"Excel file with metadata saved: {filepath}")
            
        except Exception as e:
            logger.error(f"Failed to write Excel file with metadata: {e}")
            raise
    
    def _add_metadata_sheet(self, wb: Workbook, metadata: Dict[str, Any]) -> None:
        """Add metadata sheet to workbook."""
        ws = wb.create_sheet(title='Info')
        
        # Write metadata
        row = 1
        for key, value in metadata.items():
            ws.cell(row=row, column=1, value=str(key))
            ws.cell(row=row, column=2, value=str(value))
            
            # Apply formatting
            key_cell = ws.cell(row=row, column=1)
            key_cell.font = Font(bold=True)
            key_cell.border = self.default_styles['border']
            
            value_cell = ws.cell(row=row, column=2)
            value_cell.border = self.default_styles['border']
            
            row += 1
        
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
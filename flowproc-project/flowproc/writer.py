# flowproc/writer.py
import logging
from numbers import Number
from pathlib import Path
from typing import Optional, Callable

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .config import USER_GROUP_LABELS, USER_REPLICATES
from .parsing import (
    UNKNOWN_TISSUE,
    extract_tissue,
    get_tissue_full_name,
    load_and_parse_df,
)
from .transform import map_replicates, reshape_pair

logger = logging.getLogger(__name__)

KEYWORDS = {
    "Freq. of Parent": "freq. of parent",
    "Freq. of Live": "freq. of live",
    "Median": "median",
    "Mean": "mean",
    "Count": "count",
    "GeoMean": "geomean",
    "CV": "cv",
    "SD": "sd",
    "Min": "min",
    "Max": "max",
    "Sum": "sum",
    "Mode": "mode",
    "Range": "range",
}

FRIENDLY = {
    k: v for k, v in KEYWORDS.items()
}

ALL_TIME_LABEL = "All"


# ──────────────────────────────────────────────────────────────────────────
#  Top‑level helpers
# ──────────────────────────────────────────────────────────────────────────
def _derive_replicate_count(df: pd.DataFrame) -> int:
    """Return the number of replicates in the dataframe, falling back to
    USER_REPLICATES when explicitly provided."""
    return len(USER_REPLICATES) if USER_REPLICATES else df["Replicate"].nunique()


def _clean_column_name(col) -> str:
    """Protect hashability / str operations from weird header objects."""
    return str(col).strip()


# ──────────────────────────────────────────────────────────────────────────
#  Core writer logic
# ──────────────────────────────────────────────────────────────────────────
def process_and_write_categories(
    df: pd.DataFrame,
    sid_col: str,
    wb: Workbook,
    n: int,
    time_course_mode: bool = False,
) -> None:
    """Create one pair of sheets per metric (`Count`, `Median`, …)."""

    times = (
        sorted(t for t in df["Time"].unique() if pd.notna(t))
        if "Time" in df.columns and df["Time"].notna().any()
        else [None]
    )
    groups = sorted(df["Group"].unique())

    logger.debug("Replicates per group        : %s", n)
    logger.debug("Detected time points (hours): %s", times)
    logger.debug("Detected groups             : %s", groups)

    is_timecourse = len(times) > 1 or (len(times) == 1 and times[0] is not None)
    if time_course_mode and not is_timecourse:
        logger.warning(
            "Time‑course mode selected but no time column found; falling back."
        )
        time_course_mode = False

    tissues_detected = (
        df[sid_col].apply(extract_tissue).nunique() > 1
        or any(extract_tissue(s) != UNKNOWN_TISSUE for s in df[sid_col])
    )

    # ── iterate over metric categories ───────────────────────────────────
    for cat, key_substring in KEYWORDS.items():
        raw_cols = [
            c
            for c in df.columns
            if key_substring in _clean_column_name(c).lower()
            and c
            not in {sid_col, "Well", "Group", "Animal", "Time", "Replicate"}
            and not df[c].isna().all()
        ]

        logger.debug("Metric %-15s → %s", cat, raw_cols)
        if not raw_cols:
            logger.info("No data found for metric '%s'", cat)
            continue

        friendly = FRIENDLY[cat][:31]  # max sheet name length 31

        if time_course_mode:
            _write_timecourse_sheets(
                df,
                sid_col,
                wb,
                groups,
                times,
                raw_cols,
                friendly,
                n,
            )
        else:
            _write_standard_sheets(
                df,
                sid_col,
                wb,
                groups,
                times,
                raw_cols,
                friendly,
                n,
                tissues_detected,
            )


def _write_timecourse_sheets(
    df: pd.DataFrame,
    sid_col: str,
    wb: Workbook,
    groups: list[int],
    times: list[Optional[float]],
    raw_cols: list[str],
    sheet_root: str,
    n: int,
) -> None:
    """One block per metric; replicates laid out *horizontally* in time‑course mode."""
    ws_vals = wb.create_sheet(sheet_root)
    ws_ids = wb.create_sheet(f"{sheet_root} IDs"[:31])

    col_start = 1
    for raw_col in raw_cols:
        # header line (row 1) across all group × replicate columns
        header_start = get_column_letter(col_start + 2)
        header_end = get_column_letter(col_start + 1 + len(groups) * n)
        ws_vals.merge_cells(f"{header_start}1:{header_end}1")
        ws_vals.cell(row=1, column=col_start + 2, value=raw_col).alignment = Alignment(
            horizontal="center"
        )
        ws_ids.merge_cells(f"{header_start}1:{header_end}1")
        ws_ids.cell(row=1, column=col_start + 2, value=raw_col).alignment = Alignment(
            horizontal="center"
        )

        # sub‑headers (rows 2 & 3)
        ws_vals.cell(row=2, column=col_start, value="Time")
        ws_vals.cell(row=2, column=col_start + 1, value="Group")
        ws_ids.cell(row=2, column=col_start, value="Time")
        ws_ids.cell(row=2, column=col_start + 1, value="Group")

        sub_col = col_start + 2
        for g_idx, group in enumerate(groups):
            label = (
                USER_GROUP_LABELS[g_idx]
                if g_idx < len(USER_GROUP_LABELS)
                else f"Group {group}"
            )
            c_start = get_column_letter(sub_col)
            c_end = get_column_letter(sub_col + n - 1)
            ws_vals.merge_cells(f"{c_start}2:{c_end}2")
            ws_ids.merge_cells(f"{c_start}2:{c_end}2")
            for ws in (ws_vals, ws_ids):
                ws.cell(row=2, column=sub_col, value=label).alignment = Alignment(
                    horizontal="center"
                )
            # replicates header
            for rep in range(n):
                ws_vals.cell(row=3, column=sub_col + rep, value=f"Replicate {rep+1}")
                ws_ids.cell(row=3, column=sub_col + rep, value=f"Replicate {rep+1}")
            sub_col += n

        # data rows
        row_idx = 4
        for t in times:
            t_h = int(t) if t is not None else 0
            t_m = int((t - t_h) * 60) if t is not None else 0
            time_str = f"{t_h}:{t_m:02d}" if t is not None else "0:00"

            ws_vals.cell(row=row_idx, column=col_start, value=time_str)
            ws_vals.cell(row=row_idx, column=col_start + 1, value=t_h)
            ws_ids.cell(row=row_idx, column=col_start, value=time_str)
            ws_ids.cell(row=row_idx, column=col_start + 1, value=t_h)

            sub_col = col_start + 2
            for group in groups:
                subset = df[(df["Time"] == t) & (df["Group"] == group)]
                for rep in range(1, n + 1):
                    rep_row = subset[subset["Replicate"] == rep]
                    v = rep_row[raw_col].iloc[0] if not rep_row.empty else np.nan
                    i = rep_row[sid_col].iloc[0] if not rep_row.empty else ""
                    ws_vals.cell(row=row_idx, column=sub_col, value=v)
                    ws_ids.cell(row=row_idx, column=sub_col, value=i)
                    sub_col += 1
            row_idx += 1

        col_start += 2 + len(groups) * n + 1  # space between metrics


def _write_standard_sheets(
    df: pd.DataFrame,
    sid_col: str,
    wb: Workbook,
    groups: list[int],
    times: list[Optional[float]],
    raw_cols: list[str],
    sheet_root: str,
    n: int,
    tissues_detected: bool,
) -> None:
    """Long‑format blocks *vertically* stacked by time point."""
    val_blocks, id_blocks, group_numbers, tissue_codes = [], [], [], []

    for t in times:
        chunk = df[df["Time"] == t] if t is not None else df
        if chunk.empty:
            continue
        v_blk, i_blk, t_codes, grp_nums = reshape_pair(
            chunk, sid_col, raw_cols, n, use_tissue=tissues_detected
        )
        if not v_blk:
            continue
        val_blocks.extend(v_blk)
        id_blocks.extend(i_blk)
        tissue_codes.extend(t_codes)
        group_numbers.extend(grp_nums)

    if not val_blocks:
        logger.info("No data blocks produced for sheet '%s' – skipping", sheet_root)
        return

    df_vals = pd.DataFrame(
        val_blocks,
        columns=[f"{col}_Rep{rep}" for col in raw_cols for rep in range(1, n + 1)],
    )
    df_ids = pd.DataFrame(id_blocks, columns=df_vals.columns)

    # Convert numerics safely for Excel
    df_vals = df_vals.applymap(lambda x: float(x) if isinstance(x, Number) else x)

    ws_vals = wb.create_sheet(sheet_root)
    ws_ids = wb.create_sheet(f"{sheet_root} IDs"[:31])

    # column offsets
    col_offset = 3 if tissues_detected else 2
    row_offset = 3

    # header rows 1‑2
    ws_vals.cell(row=1, column=1, value="Group")
    if tissues_detected:
        ws_vals.cell(row=1, column=2, value="Tissue")
    for ws in (ws_vals, ws_ids):
        for idx_header, col_name in enumerate(raw_cols, start=1):
            c_start = get_column_letter(col_offset + (idx_header - 1) * n)
            c_end = get_column_letter(col_offset + idx_header * n - 1)
            ws.merge_cells(f"{c_start}1:{c_end}1")
            ws.cell(row=1, column=col_offset + (idx_header - 1) * n, value=col_name).alignment = Alignment(horizontal="center")
            for rep in range(n):
                ws.cell(row=2, column=col_offset + (idx_header - 1) * n + rep, value=f"Rep {rep+1}")

    # write data rows
    for r_idx in range(len(df_vals)):
        ws_vals.cell(row=row_offset + r_idx, column=1, value=group_numbers[r_idx])
        if tissues_detected:
            ws_vals.cell(
                row=row_offset + r_idx,
                column=2,
                value=get_tissue_full_name(tissue_codes[r_idx]),
            )

    for r_idx, row in enumerate(df_vals.values, start=row_offset):
        for c_idx, val in enumerate(row, start=col_offset):
            ws_vals.cell(row=r_idx, column=c_idx, value=val)

    for r_idx, row in enumerate(df_ids.values, start=row_offset):
        for c_idx, val in enumerate(row, start=col_offset):
            ws_ids.cell(row=r_idx, column=c_idx, value=val)

    logger.info("Finished sheet '%s'", sheet_root)


# ──────────────────────────────────────────────────────────────────────────
#  Public API
# ──────────────────────────────────────────────────────────────────────────
def process_csv(
    input_file: Path,
    output_file: Path,
    time_course_mode: bool = False,
) -> None:
    """Parse one CSV and write an Excel workbook."""
    df, sid_col = load_and_parse_df(input_file)

    # replicate mapping *must* run before we count
    df = map_replicates(df)
    replicate_count = _derive_replicate_count(df)

    wb = Workbook()
    wb.remove(wb.active)

    process_and_write_categories(
        df, sid_col, wb, replicate_count, time_course_mode=time_course_mode
    )

    if not wb.sheetnames:
        wb.create_sheet("No Data")

    wb.save(output_file)
    logger.info("Saved → %s", output_file)


def process_directory(
    input_dir: Path,
    output_dir: Path,
    recursive: bool = True,
    pattern: str = "*.csv",
    status_callback: Optional[Callable[[str], None]] = None,
    time_course_mode: bool = False,
) -> None:
    """Process every CSV file under *input_dir* and write an Excel workbook
    next to it (or in *output_dir*)."""
    glob_pattern = "**/" + pattern if recursive else pattern
    csv_files = list(input_dir.glob(glob_pattern))

    if not csv_files:
        logger.warning("No CSV files found in '%s'", input_dir)
        if status_callback:
            status_callback("No CSV files found.")
        return

    for idx, f in enumerate(csv_files, 1):
        if not f.is_file():
            continue
        if status_callback:
            status_callback(f"Processing file {idx}/{len(csv_files)}: {f.name}")
        out_file = output_dir / f"{f.stem}_Processed.xlsx"
        try:
            process_csv(f, out_file, time_course_mode=time_course_mode)
        except Exception as exc:  # noqa: BLE001
            logger.exception("Error while processing '%s': %s", f, exc)
            if status_callback:
                status_callback(f"Error: {exc}")

    if status_callback:
        status_callback("Processing completed.")
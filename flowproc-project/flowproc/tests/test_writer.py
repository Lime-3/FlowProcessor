import logging
import pytest
import pandas as pd
from pathlib import Path
from typing import Optional, List, Callable
from openpyxl import load_workbook
from flowproc.writer import process_csv, process_directory
from flowproc.parsing import Constants, extract_group_animal

# Optional import for hypothesis
hypothesis = pytest.importorskip("hypothesis")
from hypothesis import given, strategies as st

@pytest.fixture
def static_did_csv(tmp_path: Path) -> Path:
    """Create DiD.csv fixture with tissue and group data."""
    content = """,DiD-A+ | Freq. of Parent (%),DiD-A+ | Median
Spleen_A1_1.1.fcs,0.73,1429
Whole Blood_B1_1.1.fcs,0.89,1103
SP_A2_1.2.fcs,0.65,1384
Whole Blood_B2_4.2.fcs,0.81,889
Mean,,
"""
    csv_file = tmp_path / "did.csv"
    csv_file.write_text(content)
    return csv_file

@pytest.fixture
def static_day4_csv(tmp_path: Path) -> Path:
    """Create AT25-AS278_Day4.csv fixture with time-course data."""
    content = """,FlowAIGoodEvents/Singlets/Live/T cells/CD4+/CD4+GFP+ | Freq. of Parent (%)
2 hour_A1_1.15.fcs,0.89
5 hour_A7_1.11.fcs,0.85
"""
    csv_file = tmp_path / "day4.csv"
    csv_file.write_text(content)
    return csv_file

@pytest.fixture
def static_simple_csv(tmp_path: Path) -> Path:
    """Create AT25-AS242.csv fixture with simple IDs."""
    content = """,Freq. of Parent
1.1.fcs,1.42
2.1.fcs,28.3
"""
    csv_file = tmp_path / "simple.csv"
    csv_file.write_text(content)
    return csv_file

@pytest.fixture
def static_multi_subpop_csv(tmp_path: Path) -> Path:
    """Create CSV with multiple subpopulations."""
    content = """,DiD-A+ | Freq. of Parent (%),DiD-A+ | Median
Spleen_A1_1.1.fcs,0.73,1429
Spleen_A2_1.2.fcs,0.65,1384
"""
    csv_file = tmp_path / "multi_subpop.csv"
    csv_file.write_text(content)
    return csv_file

@pytest.fixture
def static_invalid_csv(tmp_path: Path) -> Path:
    """Create CSV with invalid sample IDs."""
    content = """SampleID,Freq. of Parent
invalid_id,1.0
not_a_sample,2.0
"""
    csv_file = tmp_path / "invalid.csv"
    csv_file.write_text(content)
    return csv_file

@pytest.fixture
def static_duplicate_csv(tmp_path: Path) -> Path:
    """Create CSV with duplicate sample IDs."""
    content = """,Freq. of Parent
1.1.fcs,1.42
1.1.fcs,28.3
"""
    csv_file = tmp_path / "duplicate.csv"
    csv_file.write_text(content)
    return csv_file

@pytest.fixture
def static_malformed_csv(tmp_path: Path) -> Path:
    """Create CSV with non-numeric data in metric column."""
    content = """,Freq. of Parent
Spleen_A1_1.1.fcs,n/a
Spleen_A2_1.2.fcs,28.3
"""
    csv_file = tmp_path / "malformed.csv"
    csv_file.write_text(content)
    return csv_file

@pytest.fixture
def static_real_day4_csv(tmp_path: Path) -> Path:
    """Create fixture for AT25-AS278_Day4.csv with real data."""
    content = """,FlowAIGoodEvents/Singlets/Live/T cells/CD4+/CD4+GFP+ | Freq. of Parent (%),FlowAIGoodEvents/Singlets/Live/T cells/CD8+/CD8+GFP+ | Freq. of Parent (%),FlowAIGoodEvents/Singlets/Live/T cells/T cells GFP+ | Freq. of Parent (%),FlowAIGoodEvents/Singlets/Live/Non T cells/Non T cells GFP+ | Freq. of Parent (%)
2 hour_A1_1.15.fcs,0.89,0.67,0.85,1.08
2 hour_A2_1.16.fcs,0.96,0.94,1.00,0.84
5 hour_A7_1.11.fcs,0.85,0.89,0.92,1.04
5 hour_A8_1.12.fcs,0.96,0.86,1.01,0.91
"""
    csv_file = tmp_path / "real_day4.csv"
    csv_file.write_text(content)
    return csv_file

@pytest.fixture
def output_xlsx(tmp_path: Path) -> Path:
    """Provide output Excel file path."""
    return tmp_path / "output.xlsx"

@pytest.fixture
def output_dir(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """Provide shared output directory for directory tests."""
    return tmp_path_factory.mktemp("output")

def _assert_excel_structure(
    xlsx_path: Path,
    sheet_name: str,
    expected_groups: List[str],
    expected_tissues: Optional[List[str]] = None,
    expected_times: Optional[List[str]] = None,
    expected_values: Optional[List[float]] = None,
    expected_columns: Optional[List[str]] = None,
) -> None:
    """Verify Excel sheet structure and data."""
    assert xlsx_path.exists(), f"Excel file {xlsx_path} not found"
    wb = load_workbook(xlsx_path)
    assert sheet_name in wb.sheetnames, f"Sheet {sheet_name} not found"
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, skiprows=1)

    assert len(df) >= len(expected_groups), f"Expected at least {len(expected_groups)} rows, got {len(df)}"
    assert all(g in df["Group"].tolist() for g in expected_groups), f"Groups {expected_groups} not in {df['Group'].tolist()}"

    if expected_tissues:
        assert "Tissue" in df.columns, "Tissue column missing"
        assert all(t in df["Tissue"].dropna().tolist() for t in expected_tissues), f"Tissues {expected_tissues} not in {df['Tissue'].tolist()}"

    if expected_times:
        assert "Time" in df.columns, "Time column missing"
        assert df["Time"].tolist() == expected_times, f"Expected times {expected_times}, got {df['Time'].tolist()}"

    if expected_values:
        value_col = next((c for c in df.columns if c.startswith(("DiD-A+", "FlowAIGoodEvents"))), None)
        assert value_col is not None, "Value column not found"
        assert df[value_col].dropna().tolist() == expected_values, f"Expected values {expected_values}, got {df[value_col].dropna().tolist()}"

    if expected_columns:
        assert all(c in df.columns for c in expected_columns), f"Expected columns {expected_columns} not in {df.columns}"

@pytest.mark.parametrize(
    "user_groups,expected_groups,expected_values",
    [
        pytest.param([1, 4], ["Group 1", "Group 4"], [0.73, 0.89, 0.65, 0.81], id="groups_1_4"),
        pytest.param([1], ["Group 1"], [0.73, 0.89], id="single_group"),
    ],
)
def test_process_csv_did_with_tissue_and_groups(
    static_did_csv: Path, output_xlsx: Path, user_groups: List[int], expected_groups: List[str], expected_values: List[float], caplog
) -> None:
    """Test processing DiD.csv with tissue and group parsing, verifying logs."""
    caplog.set_level(logging.DEBUG)
    process_csv(static_did_csv, output_xlsx, time_course_mode=False, user_groups=user_groups)
    _assert_excel_structure(
        output_xlsx,
        sheet_name="Freq. of Parent",
        expected_groups=expected_groups,
        expected_tissues=["Spleen", "Whole Blood"],
        expected_values=expected_values,
        expected_columns=["Group", "Tissue", "DiD-A+ | Freq. of Parent (%)_Rep1"],
    )
    assert "Processing CSV: " + str(static_did_csv) in caplog.text, "Expected CSV processing log message"

def test_process_csv_timecourse(static_day4_csv: Path, output_xlsx: Path, caplog) -> None:
    """Test time-course mode with AT25-AS278_Day4.csv, verifying time logging."""
    caplog.set_level(logging.DEBUG)
    process_csv(static_day4_csv, output_xlsx, time_course_mode=True, user_groups=[1])
    _assert_excel_structure(
        output_xlsx,
        sheet_name="Freq. of Parent",
        expected_groups=["Group 1"],
        expected_times=["2:00", "5:00"],
        expected_values=[0.89, 0.85],
        expected_columns=["Group", "Time", "FlowAIGoodEvents/Singlets/Live/T cells/CD4+/CD4+GFP+ | Freq. of Parent (%)_Rep1"],
    )
    assert "Writing time value" in caplog.text, "Expected time-course logging"

def test_process_csv_simple_ids(static_simple_csv: Path, output_xlsx: Path, caplog) -> None:
    """Test processing CSV with simple IDs (AT25-AS242), verifying logs."""
    caplog.set_level(logging.DEBUG)
    process_csv(static_simple_csv, output_xlsx, time_course_mode=False, user_groups=[1, 2])
    _assert_excel_structure(
        output_xlsx,
        sheet_name="Freq. of Parent",
        expected_groups=["Group 1", "Group 2"],
        expected_values=[1.42, 28.3],
        expected_columns=["Group", "Freq. of Parent_Rep1"],
    )
    assert "Processing CSV: " + str(static_simple_csv) in caplog.text, "Expected CSV processing log message"

def test_process_multi_subpop(static_multi_subpop_csv: Path, output_xlsx: Path, caplog) -> None:
    """Test processing CSV with multiple subpopulations, verifying multiple sheets."""
    caplog.set_level(logging.DEBUG)
    process_csv(static_multi_subpop_csv, output_xlsx, time_course_mode=False, user_groups=[1])
    _assert_excel_structure(
        output_xlsx,
        sheet_name="Freq. of Parent",
        expected_groups=["Group 1"],
        expected_tissues=["Spleen"],
        expected_values=[0.73, 0.65],
        expected_columns=["Group", "Tissue", "DiD-A+ | Freq. of Parent (%)_Rep1"],
    )
    assert output_xlsx.exists()
    wb = load_workbook(output_xlsx)
    assert "Median" in wb.sheetnames, "Median sheet not created"
    df_median = pd.read_excel(output_xlsx, sheet_name="Median", skiprows=1)
    assert df_median["DiD-A+ | Median_Rep1"].dropna().tolist() == [1429, 1384], "Expected Median values"
    assert "Processing CSV: " + str(static_multi_subpop_csv) in caplog.text, "Expected CSV processing log message"

def test_process_directory(output_dir: Path, static_did_csv: Path, static_day4_csv: Path, static_simple_csv: Path, caplog) -> None:
    """Test directory processing with status callback, verifying logs."""
    caplog.set_level(logging.DEBUG)
    status_msgs: List[str] = []

    def status_callback(msg: str) -> None:
        status_msgs.append(msg)

    process_directory(
        static_did_csv.parent,
        output_dir,
        recursive=True,
        status_callback=status_callback,
        time_course_mode=False,
        user_groups=[1, 2, 4],
    )

    assert len(status_msgs) >= 4, f"Expected at least 4 status messages, got {status_msgs}"
    expected_files = [
        output_dir / "did_Processed_Grouped.xlsx",
        output_dir / "day4_Processed_Grouped.xlsx",
        output_dir / "simple_Processed_Grouped.xlsx",
    ]
    for f in expected_files:
        assert f.exists(), f"Output file {f} not created"

    _assert_excel_structure(
        output_dir / "did_Processed_Grouped.xlsx",
        sheet_name="Freq. of Parent",
        expected_groups=["Group 1", "Group 4"],
        expected_tissues=["Spleen", "Whole Blood"],
        expected_columns=["Group", "Tissue", "DiD-A+ | Freq. of Parent (%)_Rep1"],
    )
    assert "Processed 3 files" in status_msgs, "Expected completion message in status callback"
    assert "Processing directory: " in caplog.text, "Expected directory processing log message"

def test_process_csv_empty_data(static_invalid_csv: Path, output_xlsx: Path, caplog) -> None:
    """Test processing CSV with no valid data, verifying parsing warnings."""
    caplog.set_level(logging.WARNING)
    process_csv(static_invalid_csv, output_xlsx, time_course_mode=False)
    assert output_xlsx.exists(), "Excel file not created"
    wb = load_workbook(output_xlsx)
    assert "No Data" in wb.sheetnames, "No Data sheet not created"
    df = pd.read_excel(output_xlsx, sheet_name="No Data")
    assert df.empty, "No Data sheet should be empty"
    assert "No valid sample IDs found" in caplog.text, "Expected parsing warning for invalid IDs"

def test_process_csv_missing_columns(tmp_path: Path, output_xlsx: Path, caplog) -> None:
    """Test processing CSV with missing required columns, verifying no data warning."""
    caplog.set_level(logging.WARNING)
    content = """SampleID,Other
Spleen_A1_1.1.fcs,42
"""
    csv_file = tmp_path / "missing.csv"
    csv_file.write_text(content)
    process_csv(csv_file, output_xlsx, time_course_mode=False)
    assert output_xlsx.exists(), "Excel file not created"
    wb = load_workbook(output_xlsx)
    assert "No Data" in wb.sheetnames, "No Data sheet not created"
    df = pd.read_excel(output_xlsx, sheet_name="No Data")
    assert df.empty, "No Data sheet should be empty"
    assert "No valid data for" in caplog.text, "Expected no data warning"

def test_process_csv_duplicate_ids(static_duplicate_csv: Path, output_xlsx: Path, caplog) -> None:
    """Test processing CSV with duplicate sample IDs, expecting ValueError."""
    caplog.set_level(logging.WARNING)
    with pytest.raises(ValueError, match=r"Duplicate sample IDs found: \['1.1.fcs'\]"):
        process_csv(static_duplicate_csv, output_xlsx, time_course_mode=False, user_groups=[1])
    assert "Duplicate sample IDs found" in caplog.text, "Expected duplicate ID error"

def test_process_csv_malformed_data(static_malformed_csv: Path, output_xlsx: Path, caplog) -> None:
    """Test processing CSV with non-numeric data (n/a), expecting partial processing."""
    caplog.set_level(logging.WARNING)
    process_csv(static_malformed_csv, output_xlsx, time_course_mode=False, user_groups=[1])
    assert output_xlsx.exists(), "Excel file not created"
    wb = load_workbook(output_xlsx)
    assert "Freq. of Parent" in wb.sheetnames, "Freq. of Parent sheet not created"
    df = pd.read_excel(output_xlsx, sheet_name="Freq. of Parent", skiprows=1)
    assert len(df) == 1, f"Expected 1 row (non-numeric filtered out), got {len(df)}"
    assert df["Group"].tolist() == ["Group 1"], "Expected Group 1"
    assert df["Tissue"].tolist() == ["Spleen"], "Expected Spleen tissue"
    assert df["Freq. of Parent_Rep1"].dropna().tolist() == [28.3], "Expected valid value"
    assert "No columns for" not in caplog.text, "No unexpected empty column warnings"

def test_process_real_day4_csv(static_real_day4_csv: Path, output_xlsx: Path, caplog) -> None:
    """Test processing real AT25-AS278_Day4.csv in time-course mode."""
    caplog.set_level(logging.DEBUG)
    process_csv(static_real_day4_csv, output_xlsx, time_course_mode=True, user_groups=[1, 2])
    _assert_excel_structure(
        output_xlsx,
        sheet_name="Freq. of Parent",
        expected_groups=["Group 1", "Group 2"],
        expected_times=["2:00", "2:00", "5:00", "5:00"],
        expected_values=[0.89, 0.96, 0.85, 0.96],
        expected_columns=["Group", "Time", "FlowAIGoodEvents/Singlets/Live/T cells/CD4+/CD4+GFP+ | Freq. of Parent (%)_Rep1"],
    )
    assert "Processing CSV: " + str(static_real_day4_csv) in caplog.text, "Expected CSV processing log message"
    assert "Writing time value" in caplog.text, "Expected time-course logging"

@given(
    sample_id=st.text(
        min_size=1,
        max_size=30,
        alphabet=st.characters(whitelist_categories=("Lu", "Ll", "Nd"), whitelist_characters=[".", "_", " "]),
    )
)
def test_extract_group_animal_property(sample_id: str) -> None:
    """Property-based test for sample ID parsing."""
    try:
        parsed = extract_group_animal(sample_id)
        if parsed is not None:
            assert parsed.group >= 0, f"Invalid group {parsed.group} for {sample_id}"
            assert parsed.animal >= 0, f"Invalid animal {parsed.animal} for {sample_id}"
            assert isinstance(parsed.well, str), f"Well must be string, got {type(parsed.well)}"
            assert parsed.time is None or parsed.time >= 0, f"Negative time {parsed.time}"
    except ValueError:
        pass  # Expected for invalid IDs

def test_process_csv_negative_time(tmp_path: Path, output_xlsx: Path, caplog) -> None:
    """Test handling of negative time in time-course mode, verifying error logging."""
    caplog.set_level(logging.ERROR)
    content = """,FlowAIGoodEvents/Singlets/Live/T cells/CD4+/CD4+GFP+ | Freq. of Parent (%)
-2 hour_A1_1.15.fcs,0.89
"""
    csv_file = tmp_path / "negative_time.csv"
    csv_file.write_text(content)
    with pytest.raises(ValueError, match="Negative time"):
        process_csv(csv_file, output_xlsx, time_course_mode=True, user_groups=[1])
    assert "Negative time" in caplog.text, "Expected negative time error in logs"
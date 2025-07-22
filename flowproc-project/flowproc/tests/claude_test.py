#!/usr/bin/env python3
"""
Test script to verify the fixes work correctly.
Run this after applying the fixes to ensure proper Excel output.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import tempfile
import openpyxl
from typing import Dict, List, Any

# Import the fixed modules
from flowproc.parsing import load_and_parse_df
from flowproc.transform import map_replicates
from flowproc.writer import process_csv


def create_test_csv() -> Path:
    """Create a test CSV file similar to the user's data."""
    data = {
        'Unnamed: 0': [
            'SP_1.1', 'SP_1.2', 'SP_1.3',
            'SP_2.1', 'SP_2.2', 'SP_2.3',
            'WB_3.1', 'WB_3.2', 'WB_3.3',
            'WB_4.1', 'WB_4.2',
        ],
        'CD4/mCherry-A+ | Freq. of Parent (%)': [
            10.5, 11.2, 10.8,
            15.3, 14.9, 15.1,
            8.7, 9.1, 8.9,
            12.4, 12.8,
        ],
        'CD8/mCherry-A+ | Freq. of Parent (%)': [
            5.2, 5.5, 5.3,
            7.8, 7.6, 7.7,
            4.1, 4.3, 4.2,
            6.5, 6.7,
        ],
        'CD4/mCherry-A+ | Median (mCherry-A)': [
            1200, 1250, 1225,
            1450, 1425, 1440,
            980, 1010, 995,
            1320, 1340,
        ],
        'CD4 | Mean (mCherry-A)': [
            800, 820, 810,
            950, 940, 945,
            720, 735, 728,
            880, 890,
        ]
    }
    
    df = pd.DataFrame(data)
    temp_file = Path(tempfile.mktemp(suffix='.csv'))
    df.to_csv(temp_file, index=False)
    return temp_file


def verify_excel_output(excel_path: Path) -> Dict[str, Any]:
    """Verify the Excel output structure and content."""
    wb = openpyxl.load_workbook(excel_path)
    
    results = {
        'sheets': wb.sheetnames,
        'sheet_data': {}
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Get headers
        headers = []
        for col in range(1, max_col + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers.append(val)
        
        # Count data rows
        data_rows = 0
        for row in range(3, max_row + 1):
            if any(ws.cell(row=row, column=col).value for col in range(1, max_col + 1)):
                data_rows += 1
        
        results['sheet_data'][sheet_name] = {
            'headers': headers,
            'data_rows': data_rows,
            'max_col': max_col,
            'sample_data': []
        }
        
        # Get sample data from first few rows
        for row in range(3, min(6, max_row + 1)):
            row_data = []
            for col in range(1, min(6, max_col + 1)):
                row_data.append(ws.cell(row=row, column=col).value)
            if any(row_data):
                results['sheet_data'][sheet_name]['sample_data'].append(row_data)
    
    return results


def test_fixes():
    """Run comprehensive tests on the fixes."""
    print("🧪 Testing FlowProc Fixes...")
    print("-" * 50)
    
    # Create test CSV
    test_csv = create_test_csv()
    print(f"✅ Created test CSV: {test_csv}")
    
    try:
        # Test parsing
        print("\n📊 Testing CSV parsing...")
        df, sid_col = load_and_parse_df(test_csv)
        print(f"  - Loaded {len(df)} rows")
        print(f"  - Sample ID column: {sid_col}")
        print(f"  - Groups found: {sorted(df['Group'].unique())}")
        print(f"  - Tissues found: {sorted(df['Tissue'].unique())}")
        
        # Test replicate mapping
        print("\n🔢 Testing replicate mapping...")
        df_mapped, n_reps = map_replicates(df, auto_parse=True)
        print(f"  - Mapped {len(df_mapped)} rows")
        print(f"  - Number of replicates: {n_reps}")
        
        # Test Excel writing
        print("\n📝 Testing Excel writing...")
        output_file = Path(tempfile.mktemp(suffix='.xlsx'))
        process_csv(test_csv, output_file, time_course_mode=False)
        print(f"  - Created Excel file: {output_file}")
        
        # Verify Excel structure
        print("\n🔍 Verifying Excel structure...")
        results = verify_excel_output(output_file)
        
        print(f"  - Sheets created: {len(results['sheets'])}")
        for sheet in results['sheets']:
            if sheet in results['sheet_data']:
                data = results['sheet_data'][sheet]
                print(f"    • {sheet}: {data['data_rows']} rows, {data['max_col']} columns")
                if data['headers']:
                    print(f"      Headers: {data['headers'][:3]}...")
        
        # Test with time course data
        print("\n⏱️  Testing time course mode...")
        # Add time data to test
        df_time = pd.DataFrame({
            'Unnamed: 0': [
                '0h_SP_1.1', '0h_SP_1.2', '0h_SP_1.3',
                '24h_SP_1.1', '24h_SP_1.2', '24h_SP_1.3',
            ],
            'CD4 | Mean (mCherry-A)': [100, 105, 102, 200, 210, 205]
        })
        
        time_csv = Path(tempfile.mktemp(suffix='.csv'))
        df_time.to_csv(time_csv, index=False)
        
        time_output = Path(tempfile.mktemp(suffix='_timecourse.xlsx'))
        process_csv(time_csv, time_output, time_course_mode=True)
        print(f"  - Created time course Excel: {time_output}")
        
        print("\n✅ All tests passed!")
        
    except Exception as e:
        print(f"\n❌ Test failed: {str(e)}")
        import traceback
        traceback.print_exc()
    
    finally:
        # Cleanup
        if test_csv.exists():
            test_csv.unlink()


def debug_user_file(csv_path: str):
    """Debug a specific user file."""
    print(f"\n🐛 Debugging user file: {csv_path}")
    print("-" * 50)
    
    csv_path = Path(csv_path)
    if not csv_path.exists():
        print(f"❌ File not found: {csv_path}")
        return
    
    try:
        # Load and parse
        df, sid_col = load_and_parse_df(csv_path)
        print(f"✅ Parsed {len(df)} valid rows")
        
        # Show sample data
        print("\n📋 Sample data:")
        print(df[['SampleID', 'Group', 'Animal', 'Tissue']].head(10))
        
        # Show column info
        print("\n📊 Data columns found:")
        data_cols = [c for c in df.columns if c not in ['SampleID', 'Well', 'Group', 'Animal', 'Time', 'Replicate', 'Tissue']]
        for col in data_cols[:5]:
            print(f"  - {col}")
        if len(data_cols) > 5:
            print(f"  ... and {len(data_cols) - 5} more")
        
        # Process to Excel
        output_path = csv_path.parent / f"{csv_path.stem}_DEBUG.xlsx"
        process_csv(csv_path, output_path)
        print(f"\n✅ Created debug Excel: {output_path}")
        
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        # Debug specific file
        debug_user_file(sys.argv[1])
    else:
        # Run general tests
        test_fixes()
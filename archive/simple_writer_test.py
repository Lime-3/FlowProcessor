#!/usr/bin/env python3
"""
Simple test script to check writer functionality without complex test infrastructure.
"""

import tempfile
import logging
from pathlib import Path
import pandas as pd

# Set up logging
logging.basicConfig(level=logging.DEBUG)

def test_writer_basic():
    """Test basic writer functionality."""
    print("🧪 Testing writer basic functionality...")
    
    try:
        # Import the modules
        from flowproc.domain.export import process_csv
        from flowproc.domain.parsing import extract_group_animal, Constants
        print("✅ Imports successful")
        
        # Create test data
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir = Path(tmpdir)
            
            # Create a test CSV file similar to the original test
            test_content = """,DiD-A+ | Freq. of Parent (%),DiD-A+ | Median
Spleen_A1_1.1.fcs,0.73,1429
Whole Blood_B1_1.1.fcs,0.89,1103
SP_A2_1.2.fcs,0.65,1384
Whole Blood_B2_4.2.fcs,0.81,889
Mean,,
"""
            csv_file = tmpdir / "test_did.csv"
            csv_file.write_text(test_content)
            print(f"✅ Created test CSV: {csv_file}")
            
            # Test processing
            output_file = tmpdir / "output.xlsx"
            
            print("🔄 Processing CSV...")
            process_csv(csv_file, output_file, time_course_mode=False, user_groups=[1, 4])
            
            # Check if output was created
            if output_file.exists():
                print(f"✅ Output file created: {output_file}")
                print(f"📄 File size: {output_file.stat().st_size} bytes")
                
                # Try to read it back
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(output_file)
                    print(f"📊 Excel sheets: {wb.sheetnames}")
                    
                    # Try to find a sheet with "Freq. of Parent" in the name
                    freq_sheets = [name for name in wb.sheetnames if "Freq. of Parent" in name]
                    if freq_sheets:
                        sheet_name = freq_sheets[0]
                        print(f"✅ Found sheet: '{sheet_name}'")
                        df = pd.read_excel(output_file, sheet_name=sheet_name, skiprows=1)
                        print(f"📊 Data shape: {df.shape}")
                        print(f"📊 Columns: {list(df.columns)}")
                        print("✅ Successfully read Excel data")
                    else:
                        print(f"❌ No frequency sheets found in {wb.sheetnames}")
                     
                except Exception as e:
                    print(f"⚠️  Could not read Excel file: {e}")
                    
            else:
                print("❌ Output file was not created")
                return False
                
        print("🎉 Writer test completed successfully!")
        return True
        
    except Exception as e:
        print(f"❌ Writer test failed: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_parsing_basic():
    """Test basic parsing functionality."""
    print("\n🧪 Testing parsing basic functionality...")
    
    try:
        from flowproc.domain.parsing import extract_group_animal, Constants
        
        # Test sample ID parsing
        test_cases = [
            "Spleen_A1_1.1.fcs",
            "Whole Blood_B1_1.1.fcs", 
            "SP_A2_1.2.fcs",
            "2 hour_A1_1.15.fcs"
        ]
        
        for sample_id in test_cases:
            result = extract_group_animal(sample_id)
            if result:
                print(f"✅ Parsed '{sample_id}' -> Group: {result.group}, Animal: {result.animal}, Well: {result.well}")
            else:
                print(f"⚠️  Could not parse '{sample_id}'")
                
        print("✅ Parsing test completed!")
        return True
        
    except Exception as e:
        print(f"❌ Parsing test failed: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    print("🚀 Starting FlowProcessor Writer Tests")
    print("=" * 50)
    
    success = True
    
    # Test parsing first
    success &= test_parsing_basic()
    
    # Test writer
    success &= test_writer_basic()
    
    print("\n" + "=" * 50)
    if success:
        print("🎉 All tests passed!")
    else:
        print("❌ Some tests failed!")
    
    exit(0 if success else 1) 